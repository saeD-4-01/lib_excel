from lib_logger import Logger
import openpyxl

class Excel:

    def __init__(self, path, logger:Logger):
        self.path = path
        self.logger = logger
        try :
            self.wb = openpyxl.load_workbook(self.path)
        except Exception as e :
            self.logger.error("Impossible d'ouvrir le fichier excel : " + e)

    def table_convert(self, page=None, nb_column=None, nb_line=None):
        sheet = self.open_page(page)
        tableau = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # Gestion du nombre de lignes
            if nb_line is not None and i >= nb_line:
                break
            # Gestion du nombre de colonnes
            if nb_column is not None:
                row = row[:nb_column]
            tableau.append(list(row))

        return tableau

    def get_headers(self, page=None, nb_column=None):
        return self.get_line(1)

    def get_line(self, line, page=None, nb_column=None):
        sheet = self.open_page(page)
        if sheet is False:
            return None  # ou raise une exception selon ton choix
        # On récupère toutes les lignes de la feuille
        all_rows = list(sheet.iter_rows(values_only=True))
        # Vérification que la ligne demandée existe
        if line < 1 or line > len(all_rows):
            self.logger.error(f"La ligne {line} n'existe pas dans la feuille.")
            return None
        # On récupère la ligne demandée (attention, Python indexe à partir de 0)
        row = all_rows[line - 1]
        # Si un nombre de colonnes est spécifié, on tronque la ligne
        if nb_column is not None:
            row = row[:nb_column]
        return list(row)

    def write_cell(self, row, column, value, page=None, save=True):
        """
        Écrit une valeur dans une cellule spécifique de la feuille Excel.

        :param row: numéro de ligne (1-indexé, comme dans Excel)
        :param column: numéro de colonne (1-indexé, comme dans Excel)
        :param value: valeur à écrire dans la cellule
        :param page: nom de la feuille (optionnel, active par défaut)
        :param save: booléen, si True, sauvegarde le fichier après écriture
        :return: True si succès, False sinon
        """
        sheet = self.open_page(page)
        if sheet is False:
            self.logger.error("Écriture impossible : feuille non trouvée.")
            return False

        try:
            sheet.cell(row=row, column=column, value=value)
            if save:
                self.wb.save(self.path)
            return True
        except Exception as e:
            self.logger.error(f"Erreur lors de l'écriture dans la cellule : {e}")
            return False

    def open_page(self, page):
        try:
            if page is None:
                return self.wb.active
            return self.wb[page]
        except Exception as e :
            self.logger.error("Impossible d'ouvrir la page spécifiée : " + e)
            return False